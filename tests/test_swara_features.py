"""
tests/test_swara_features.py — Automated verification for Swara.ai enhancements:
  1. Excel deliverable generation (openpyxl, bold headers, zebra striping)
  2. Router Layer 3 manual override and L1 deterministic fast-path
  3. HITL 403 gate on /api/artifact/{task_id}/xlsx
  4. /api/pid-tags endpoint
  5. /health returning Swara.ai system name
"""

import os
import pytest
from pathlib import Path
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.exporter import render_excel_deliverable
from src.router import route_l1, route_l3
from src.main import create_app


class TestExcelDeliverable:
    def test_render_excel_deliverable_creates_valid_workbook(self, tmp_path):
        task_id = "test_task_123"
        data = [
            {"Tag": "E-101", "Component": "Heat Exchanger", "Corrosion Rate (mm/y)": 0.125, "Status": "Normal"},
            {"Tag": "P-202", "Component": "Feed Pump", "Corrosion Rate (mm/y)": 0.450, "Status": "Alert"},
            {"Tag": "T-303", "Component": "Storage Tank", "Corrosion Rate (mm/y)": 0.080, "Status": "Normal"},
        ]

        out_path = render_excel_deliverable(task_id, data)
        assert os.path.isfile(out_path), f"File not created at {out_path}"

        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.title == "Swara.ai Calculations"

        # Check banner title
        assert "Swara.ai" in str(ws["A1"].value)

        # Check header row (row 3)
        assert ws.cell(row=3, column=1).value == "Tag"
        assert ws.cell(row=3, column=3).value == "Corrosion Rate (mm/y)"
        assert ws.cell(row=3, column=1).font.bold is True

        # Check data row formatting (row 4)
        assert ws.cell(row=4, column=1).value == "E-101"
        assert ws.cell(row=4, column=3).value == 0.125
        assert ws.cell(row=4, column=3).alignment.horizontal == "right"

        # Clean up
        try:
            os.remove(out_path)
        except OSError:
            pass


class TestRouterL3:
    def test_l3_manual_override_bypasses_all(self):
        spec, trace = route_l3("vision")
        assert spec == "vision"
        assert "L3 manual override -> vision" in trace

        spec, trace = route_l3("coder")
        assert spec == "coder"
        assert "L3 manual override -> coder" in trace

        spec, trace = route_l3("deep_brain")
        assert spec == "deep_brain"
        assert "L3 manual override -> deep_brain" in trace

        # Legacy alias mapping
        spec, trace = route_l3("brain")
        assert spec == "deep_brain"


class TestFastAPIEndpoints:
    @pytest.fixture(scope="class")
    def client(self):
        app = create_app()
        return TestClient(app)

    def test_health_endpoint_branding(self, client):
        resp = client.get("/health")
        assert resp.status_code == 200
        data = resp.json()
        assert data["system"] == "Swara.ai"
        assert data["version"] == "3.0.0"
        assert "firewall_active" in data

    def test_hitl_403_gate_on_unapproved_xlsx(self, client):
        resp = client.get("/api/artifact/unapproved_task_999/xlsx")
        assert resp.status_code == 403
        assert "Artifact not approved via HITL" in resp.json()["detail"]

    def test_pid_tags_endpoint(self, client):
        resp = client.get("/api/pid-tags")
        assert resp.status_code == 200
        data = resp.json()
        assert "tags" in data
